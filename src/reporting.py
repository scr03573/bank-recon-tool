"""
Reconciliation reporting and export functionality.

Generates:
- Excel reports with multiple worksheets
- Summary dashboards
- Exception reports
- Audit trails
"""
from dataclasses import asdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Dict, Any, Optional
import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
from jinja2 import Template

from .models import (
    ReconciliationMatch, ReconciliationException, ReconciliationSummary,
    BankTransaction, APTransaction, MatchStatus, ExceptionType
)
from .economic_context import EconomicSnapshot


class ReportGenerator:
    """Generates reconciliation reports in various formats."""

    # Excel styling
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    MATCHED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    EXCEPTION_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, output_dir: Optional[Path] = None):
        self.output_dir = output_dir or Path("./reports")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_excel_report(
        self,
        summary: ReconciliationSummary,
        matches: List[ReconciliationMatch],
        exceptions: List[ReconciliationException],
        economic_snapshot: Optional[EconomicSnapshot] = None,
        filename: Optional[str] = None
    ) -> Path:
        """Generate comprehensive Excel reconciliation report."""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_summary_sheet(wb, summary, economic_snapshot)
        self._create_matches_sheet(wb, matches)
        self._create_exceptions_sheet(wb, exceptions)
        self._create_unmatched_bank_sheet(wb, matches, summary)
        self._create_unmatched_ap_sheet(wb, matches, summary)

        # Save workbook
        if not filename:
            filename = f"reconciliation_{summary.period_end.strftime('%Y%m%d')}.xlsx"

        output_path = self.output_dir / filename
        wb.save(output_path)

        return output_path

    def _create_summary_sheet(
        self,
        wb: Workbook,
        summary: ReconciliationSummary,
        economic: Optional[EconomicSnapshot]
    ):
        """Create summary dashboard sheet."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "Bank Reconciliation Summary"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        # Period info
        ws["A3"] = "Reconciliation Period:"
        ws["B3"] = f"{summary.period_start} to {summary.period_end}"
        ws["A4"] = "Bank Account:"
        ws["B4"] = summary.bank_account_id
        ws["A5"] = "Run Date:"
        ws["B5"] = summary.run_date.strftime("%Y-%m-%d %H:%M:%S")

        # Transaction counts
        ws["A7"] = "Transaction Summary"
        ws["A7"].font = Font(bold=True, size=12)

        count_data = [
            ("Bank Transactions", summary.total_bank_transactions),
            ("AP Transactions", summary.total_ap_transactions),
            ("Matched", summary.matched_count),
            ("Partial Matches", summary.partial_match_count),
            ("Unmatched (Bank)", summary.unmatched_bank_count),
            ("Unmatched (AP)", summary.unmatched_ap_count),
            ("Exceptions", summary.exception_count),
        ]

        for i, (label, value) in enumerate(count_data, start=8):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value

        # Amounts
        ws["A16"] = "Amount Summary"
        ws["A16"].font = Font(bold=True, size=12)

        amount_data = [
            ("Total Bank Amount", summary.total_bank_amount),
            ("Total AP Amount", summary.total_ap_amount),
            ("Matched Amount", summary.matched_amount),
            ("Unreconciled Amount", summary.unreconciled_amount),
        ]

        for i, (label, value) in enumerate(amount_data, start=17):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = float(value)
            ws[f"B{i}"].number_format = '$#,##0.00'

        # Performance metrics
        ws["A23"] = "Performance Metrics"
        ws["A23"].font = Font(bold=True, size=12)

        ws["A24"] = "Auto-Match Rate:"
        ws["B24"] = f"{summary.auto_match_rate:.1%}"
        ws["A25"] = "Processing Time:"
        ws["B25"] = f"{summary.processing_time_seconds:.2f} seconds"

        # Economic context
        if economic:
            ws["D3"] = "Economic Context"
            ws["D3"].font = Font(bold=True, size=12)

            econ_data = [
                ("Fed Funds Rate", f"{economic.fed_funds_rate:.2f}%" if economic.fed_funds_rate else "N/A"),
                ("10Y Treasury", f"{economic.treasury_10y:.2f}%" if economic.treasury_10y else "N/A"),
                ("VIX", f"{economic.vix:.1f}" if economic.vix else "N/A"),
                ("S&P 500", f"${economic.sp500_price:,.0f}" if economic.sp500_price else "N/A"),
            ]

            for i, (label, value) in enumerate(econ_data, start=4):
                ws[f"D{i}"] = label
                ws[f"E{i}"] = value

        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 15

    def _create_matches_sheet(self, wb: Workbook, matches: List[ReconciliationMatch]):
        """Create matched transactions sheet."""
        ws = wb.create_sheet("Matched Transactions")

        headers = [
            "Match ID", "Status", "Confidence", "Bank Date", "Bank Amount",
            "Bank Description", "AP Vendor", "AP Amount", "Variance",
            "Match Reasons", "Check Number"
        ]

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER

        # Write data
        for row_num, match in enumerate(matches, start=2):
            bank_tx = match.bank_transaction
            ap_vendors = ", ".join([ap.vendor_name for ap in match.ap_transactions])
            ap_total = sum(ap.paid_amount for ap in match.ap_transactions)

            row_data = [
                match.id[:8],
                match.match_status.value,
                f"{match.confidence_score:.1%}",
                bank_tx.transaction_date if bank_tx else "",
                float(abs(bank_tx.amount)) if bank_tx else 0,
                bank_tx.description[:50] if bank_tx else "",
                ap_vendors[:40],
                float(ap_total),
                float(match.variance),
                "; ".join(match.match_reasons[:3]),
                bank_tx.check_number if bank_tx else ""
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = self.BORDER

                # Apply conditional formatting
                if col == 2:  # Status column
                    if match.match_status == MatchStatus.MATCHED:
                        cell.fill = self.MATCHED_FILL
                    elif match.match_status == MatchStatus.MANUAL_REVIEW:
                        cell.fill = self.WARNING_FILL

            # Format currency columns
            ws.cell(row=row_num, column=5).number_format = '$#,##0.00'
            ws.cell(row=row_num, column=8).number_format = '$#,##0.00'
            ws.cell(row=row_num, column=9).number_format = '$#,##0.00'

        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _create_exceptions_sheet(self, wb: Workbook, exceptions: List[ReconciliationException]):
        """Create exceptions report sheet."""
        ws = wb.create_sheet("Exceptions")

        headers = [
            "Exception ID", "Type", "Severity", "Description",
            "Suggested Action", "Bank Ref", "AP Ref", "Created"
        ]

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.BORDER

        # Write data
        for row_num, exc in enumerate(exceptions, start=2):
            bank_ref = exc.bank_transaction.reference_number if exc.bank_transaction else ""
            ap_ref = exc.ap_transaction.record_number if exc.ap_transaction else ""

            row_data = [
                exc.id[:8],
                exc.exception_type.value,
                exc.severity,
                exc.description[:60],
                exc.suggested_action[:40],
                bank_ref,
                ap_ref,
                exc.created_at.strftime("%Y-%m-%d %H:%M")
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = self.BORDER

                # Color by severity
                if col == 3:
                    if exc.severity == "critical":
                        cell.fill = self.EXCEPTION_FILL
                    elif exc.severity == "high":
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    elif exc.severity == "medium":
                        cell.fill = self.WARNING_FILL

        # Set column widths
        widths = [12, 20, 12, 50, 40, 15, 15, 18]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _create_unmatched_bank_sheet(
        self,
        wb: Workbook,
        matches: List[ReconciliationMatch],
        summary: ReconciliationSummary
    ):
        """Create sheet for unmatched bank transactions."""
        ws = wb.create_sheet("Unmatched Bank")

        headers = ["Date", "Amount", "Description", "Type", "Reference", "Check #"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        # Get matched bank IDs
        matched_ids = {m.bank_transaction.id for m in matches if m.bank_transaction}

        # This would need access to all bank transactions
        # For now, add a note
        ws["A3"] = "Note: Unmatched bank transactions are tracked in the Exceptions sheet"

    def _create_unmatched_ap_sheet(
        self,
        wb: Workbook,
        matches: List[ReconciliationMatch],
        summary: ReconciliationSummary
    ):
        """Create sheet for unmatched AP transactions."""
        ws = wb.create_sheet("Unmatched AP")

        headers = ["Vendor", "Amount", "Payment Date", "Bill #", "Check #", "State"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        ws["A3"] = "Note: Unmatched AP transactions are tracked in the Exceptions sheet"

    def generate_json_report(
        self,
        summary: ReconciliationSummary,
        matches: List[ReconciliationMatch],
        exceptions: List[ReconciliationException],
        filename: Optional[str] = None
    ) -> Path:
        """Generate JSON report for API consumption or further processing."""
        report_data = {
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "period_start": summary.period_start.isoformat() if summary.period_start else None,
                "period_end": summary.period_end.isoformat() if summary.period_end else None,
                "bank_account_id": summary.bank_account_id,
                "total_bank_transactions": summary.total_bank_transactions,
                "total_ap_transactions": summary.total_ap_transactions,
                "matched_count": summary.matched_count,
                "exception_count": summary.exception_count,
                "auto_match_rate": summary.auto_match_rate,
                "total_bank_amount": str(summary.total_bank_amount),
                "total_ap_amount": str(summary.total_ap_amount),
                "unreconciled_amount": str(summary.unreconciled_amount),
            },
            "matches": [self._match_to_dict(m) for m in matches],
            "exceptions": [self._exception_to_dict(e) for e in exceptions],
        }

        if not filename:
            filename = f"reconciliation_{summary.period_end.strftime('%Y%m%d')}.json"

        output_path = self.output_dir / filename
        with open(output_path, 'w') as f:
            json.dump(report_data, f, indent=2, default=str)

        return output_path

    def _match_to_dict(self, match: ReconciliationMatch) -> Dict[str, Any]:
        """Convert match to dictionary."""
        return {
            "id": match.id,
            "status": match.match_status.value,
            "confidence": match.confidence_score,
            "bank_transaction": {
                "date": str(match.bank_transaction.transaction_date) if match.bank_transaction else None,
                "amount": str(abs(match.bank_transaction.amount)) if match.bank_transaction else None,
                "description": match.bank_transaction.description if match.bank_transaction else None,
            },
            "ap_transactions": [
                {
                    "vendor": ap.vendor_name,
                    "amount": str(ap.paid_amount),
                    "date": str(ap.payment_date) if ap.payment_date else None,
                }
                for ap in match.ap_transactions
            ],
            "variance": str(match.variance),
            "reasons": match.match_reasons,
        }

    def _exception_to_dict(self, exc: ReconciliationException) -> Dict[str, Any]:
        """Convert exception to dictionary."""
        return {
            "id": exc.id,
            "type": exc.exception_type.value,
            "severity": exc.severity,
            "description": exc.description,
            "suggested_action": exc.suggested_action,
            "resolved": exc.resolved,
        }

    def generate_html_report(
        self,
        summary: ReconciliationSummary,
        matches: List[ReconciliationMatch],
        exceptions: List[ReconciliationException],
        economic: Optional[EconomicSnapshot] = None,
        filename: Optional[str] = None
    ) -> Path:
        """Generate HTML report for web viewing."""
        template = Template(HTML_REPORT_TEMPLATE)

        html_content = template.render(
            summary=summary,
            matches=matches,
            exceptions=exceptions,
            economic=economic,
            generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )

        if not filename:
            filename = f"reconciliation_{summary.period_end.strftime('%Y%m%d')}.html"

        output_path = self.output_dir / filename
        with open(output_path, 'w') as f:
            f.write(html_content)

        return output_path


# HTML Report Template
HTML_REPORT_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Bank Reconciliation Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        h1 { color: #333; border-bottom: 2px solid #366092; padding-bottom: 10px; }
        h2 { color: #366092; margin-top: 30px; }
        .summary-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }
        .summary-card { background: #f8f9fa; padding: 15px; border-radius: 8px; border-left: 4px solid #366092; }
        .summary-card h3 { margin: 0 0 10px 0; color: #666; font-size: 14px; }
        .summary-card .value { font-size: 24px; font-weight: bold; color: #333; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background: #366092; color: white; }
        tr:hover { background: #f5f5f5; }
        .status-matched { background: #c6efce; color: #006100; padding: 4px 8px; border-radius: 4px; }
        .status-exception { background: #ffc7ce; color: #9c0006; padding: 4px 8px; border-radius: 4px; }
        .status-review { background: #ffeb9c; color: #9c5700; padding: 4px 8px; border-radius: 4px; }
        .severity-high { color: #d63031; font-weight: bold; }
        .severity-medium { color: #fdcb6e; }
        .severity-low { color: #00b894; }
        .economic-banner { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 15px; border-radius: 8px; margin: 20px 0; }
        .economic-items { display: flex; gap: 30px; flex-wrap: wrap; }
        .economic-item { text-align: center; }
        .economic-item .label { font-size: 12px; opacity: 0.8; }
        .economic-item .value { font-size: 18px; font-weight: bold; }
        .footer { margin-top: 40px; text-align: center; color: #666; font-size: 12px; }
    </style>
</head>
<body>
    <div class="container">
        <h1>Bank Reconciliation Report</h1>
        <p>Period: {{ summary.period_start }} to {{ summary.period_end }} | Generated: {{ generated_at }}</p>

        {% if economic %}
        <div class="economic-banner">
            <strong>Economic Context</strong>
            <div class="economic-items">
                {% if economic.fed_funds_rate %}<div class="economic-item"><div class="label">Fed Funds</div><div class="value">{{ "%.2f"|format(economic.fed_funds_rate) }}%</div></div>{% endif %}
                {% if economic.treasury_10y %}<div class="economic-item"><div class="label">10Y Treasury</div><div class="value">{{ "%.2f"|format(economic.treasury_10y) }}%</div></div>{% endif %}
                {% if economic.vix %}<div class="economic-item"><div class="label">VIX</div><div class="value">{{ "%.1f"|format(economic.vix) }}</div></div>{% endif %}
            </div>
        </div>
        {% endif %}

        <div class="summary-grid">
            <div class="summary-card">
                <h3>Bank Transactions</h3>
                <div class="value">{{ summary.total_bank_transactions }}</div>
            </div>
            <div class="summary-card">
                <h3>AP Transactions</h3>
                <div class="value">{{ summary.total_ap_transactions }}</div>
            </div>
            <div class="summary-card">
                <h3>Matched</h3>
                <div class="value">{{ summary.matched_count }}</div>
            </div>
            <div class="summary-card">
                <h3>Exceptions</h3>
                <div class="value">{{ summary.exception_count }}</div>
            </div>
            <div class="summary-card">
                <h3>Auto-Match Rate</h3>
                <div class="value">{{ "%.1f"|format(summary.auto_match_rate * 100) }}%</div>
            </div>
            <div class="summary-card">
                <h3>Unreconciled</h3>
                <div class="value">${{ "%.2f"|format(summary.unreconciled_amount|float) }}</div>
            </div>
        </div>

        <h2>Matched Transactions</h2>
        <table>
            <thead>
                <tr>
                    <th>Status</th>
                    <th>Confidence</th>
                    <th>Bank Date</th>
                    <th>Amount</th>
                    <th>Vendor</th>
                    <th>Match Reasons</th>
                </tr>
            </thead>
            <tbody>
                {% for match in matches %}
                {% if loop.index <= 50 %}
                <tr>
                    <td>
                        {% if match.match_status.value == 'matched' %}
                        <span class="status-matched">Matched</span>
                        {% elif match.match_status.value == 'manual_review' %}
                        <span class="status-review">Review</span>
                        {% else %}
                        <span class="status-exception">{{ match.match_status.value }}</span>
                        {% endif %}
                    </td>
                    <td>{{ "%.0f"|format(match.confidence_score * 100) }}%</td>
                    <td>{{ match.bank_transaction.transaction_date if match.bank_transaction else '-' }}</td>
                    <td>${{ "%.2f"|format(match.bank_transaction.amount|abs|float) if match.bank_transaction else '0' }}</td>
                    <td>{% if match.ap_transactions %}{{ match.ap_transactions[0].vendor_name|truncate(30) }}{% else %}-{% endif %}</td>
                    <td>{{ match.match_reasons|join(', ')|truncate(50) }}</td>
                </tr>
                {% endif %}
                {% endfor %}
            </tbody>
        </table>
        {% if matches|length > 50 %}<p><em>Showing 50 of {{ matches|length }} matches</em></p>{% endif %}

        <h2>Exceptions ({{ exceptions|length }})</h2>
        <table>
            <thead>
                <tr>
                    <th>Type</th>
                    <th>Severity</th>
                    <th>Description</th>
                    <th>Suggested Action</th>
                </tr>
            </thead>
            <tbody>
                {% for exc in exceptions %}
                {% if loop.index <= 30 %}
                <tr>
                    <td>{{ exc.exception_type.value }}</td>
                    <td class="severity-{{ exc.severity }}">{{ exc.severity }}</td>
                    <td>{{ exc.description|truncate(60) }}</td>
                    <td>{{ exc.suggested_action|truncate(40) }}</td>
                </tr>
                {% endif %}
                {% endfor %}
            </tbody>
        </table>
        {% if exceptions|length > 30 %}<p><em>Showing 30 of {{ exceptions|length }} exceptions</em></p>{% endif %}

        <div class="footer">
            Generated by Bank Reconciliation Tool | {{ generated_at }}
        </div>
    </div>
</body>
</html>
"""
